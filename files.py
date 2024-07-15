import os
import pytz
import json
import time
import datetime
import openpyxl
import numpy as np


def read_queries(filename="./文章测试query.xlsx", sheet_name=None, num_queries_per_sheet=150, num_queries_to_read_per_sheet=15):
    queries = []
    workbook = openpyxl.load_workbook(filename=filename, read_only=True, data_only=True, keep_links=False)
    for sheet in workbook.worksheets:
        if sheet_name is not None and sheet.title!=sheet_name:
            continue
        for i in range(0, num_queries_per_sheet, num_queries_per_sheet//num_queries_to_read_per_sheet):
            queries.append(sheet[f'A{i+2}'].value)
    return queries


def read_cached_answers(cached_file="./query2answers.json"):
    if os.path.exists(cached_file):
        with open(cached_file, "r", encoding="utf-8") as f:
            query2answers = json.loads(f.read())
    else:
        query2answers = {}
    return query2answers


def save_cached_answers(query2answers, cached_file="./query2answers.json"):
    with open(cached_file, "w", encoding="utf-8") as f:
        f.write(json.dumps(query2answers, ensure_ascii=False, indent=4))
    return


def update_cached_answers_with_results_in_excel(excel_file="./outputs-2023-08-30/results.18-30-22.xlsx"):
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    query2answers = read_cached_answers()
    for row in sheet.iter_rows(min_row=2, max_row=101, min_col=1, max_col=2, values_only=True):
        query, references = row
        if references is None:
            continue
        references = references.split("\n")
        assert all([reference.startswith(f"{i+1}. ") for i,reference in enumerate(references)])
        assert all(["：" in reference for reference in references])
        title2abstract = {reference[len(f"{i+1}. "):reference.index("：")]:reference[reference.index("：")+1:]
                          for i, reference in enumerate(references)}
        title2abstract.update(query2answers.get(query, {}))
        query2answers[query] = title2abstract
    save_cached_answers(query2answers)


def save_results(results):
    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet(title="results")
    titles = list(results.keys())
    sheet.append(titles)
    row_lengths = [[len(title) for title in titles]]
    for row in zip(*[v for k,v in results.items()]):
        sheet.append([_ if _ is None else str(_) for _ in row])
        row_lengths.append([max([len(line) for line in str(content).split("\n")]) for content in row])
    column_width = (np.array(row_lengths).mean(axis=0)*1.2).tolist()
    for column,width in enumerate(column_width):
        sheet.column_dimensions[chr(ord('A')+column)].width = width

    timezone = pytz.timezone('Asia/Shanghai')
    timestamp2string = lambda timestamp: datetime.datetime.fromtimestamp(
        timestamp).astimezone(timezone).strftime('%Y-%m-%d.%H-%M-%S')
    timestamp = timestamp2string(time.time())
    save_dir = f"outputs-{timestamp.split('.')[0]}"
    if not os.path.exists(save_dir):
        os.mkdir(save_dir)
    workbook.save(filename=f"{save_dir}/results.{timestamp.split('.')[1]}.xlsx")


if __name__ == "__main__":
    update_cached_answers_with_results_in_excel()